import json
import os
import base64
import shutil
import time
import sys
from datetime import datetime
from openpyxl import load_workbook
from google import genai
from concurrent.futures import ThreadPoolExecutor, as_completed
import threading
import config

TRACK_FILE = "batch_tracking.json"
OUTPUT_DIR = r"C:\ShareFolder\ batch_downloaded_images"
MAX_WORKERS = 8  # parallel image saves + Excel updates

client = genai.Client(api_key=config.GEMINI_API_KEY)
os.makedirs(OUTPUT_DIR, exist_ok=True)

wb_lock = threading.Lock()


# ==============================
# Progress spinner for long downloads
# ==============================
def show_spinner(stop_event, message="Downloading"):
    """Show animated dots while downloading."""
    spinner = [".", "..", "...", ""]
    idx = 0
    while not stop_event.is_set():
        sys.stdout.write(f"\r    {message}{spinner[idx % len(spinner)]}   ")
        sys.stdout.flush()
        time.sleep(0.5)
        idx += 1
    sys.stdout.write("\r")
    sys.stdout.flush()


# ==============================
# Load / Save Tracking
# ==============================
def load_tracking():
    if not os.path.exists(TRACK_FILE):
        return []
    with open(TRACK_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def save_tracking(data):
    with open(TRACK_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)


# ==============================
# Save Image (no overwrite)
# ==============================
def save_image_safely(image_key, image_bytes):
    path = os.path.join(OUTPUT_DIR, f"{image_key}.jpg")
    counter = 1
    while os.path.exists(path):
        path = os.path.join(OUTPUT_DIR, f"{image_key}-{counter}.jpg")
        counter += 1
    with open(path, "wb") as f:
        f.write(image_bytes)
    return path


# ==============================
# Process a single result line
# ==============================
def process_result_line(item, ws, headers, imagename_to_row):
    """
    Decode one JSONL line: save image + mark Excel row as Done.
    Returns (image_key, status_str).
    """
    image_key = item.get("key")
    if not image_key:
        return None, "no_key"

    if "error" in item:
        return image_key, "api_error"

    response = item.get("response", {})
    candidates = response.get("candidates", [])
    if not candidates:
        return image_key, "no_candidates"

    parts = candidates[0].get("content", {}).get("parts", [])
    saved = False
    for part in parts:
        if "inlineData" in part:
            image_bytes = base64.b64decode(part["inlineData"]["data"])
            save_image_safely(image_key, image_bytes)
            saved = True
            break

    if not saved:
        return image_key, "no_image"

    # Update Excel using lookup dict (fast)
    image_key_clean = str(image_key).strip().lower()
    with wb_lock:
        row = imagename_to_row.get(image_key_clean)
        if row:
            ws.cell(row=row, column=headers["status"]).value = "Done"

    return image_key, "saved"


# ==============================
# Find result file from job object
# ==============================
def find_result_file(job):
    """Extract the result file reference from a batch job."""
    # Try direct attribute first
    try:
        if hasattr(job, "dest") and job.dest:
            name = getattr(job.dest, "file_name", None) or getattr(job.dest, "name", None)
            if name:
                return name
    except Exception:
        pass

    # Try result attribute
    try:
        if getattr(job, "result", None):
            rf = getattr(job.result, "file", None) or getattr(job.result, "uri", None) or getattr(job.result, "name", None)
            if rf:
                return rf
    except Exception:
        pass

    # Walk the model dump looking for file-like strings
    try:
        dump = job.model_dump() if hasattr(job, "model_dump") else {}
    except Exception:
        dump = {}

    def find_candidate(d):
        if isinstance(d, dict):
            for k, v in d.items():
                if isinstance(v, str):
                    s = v.strip()
                    if s.startswith("files/") or s.endswith(".jsonl") or s.endswith(".json"):
                        return s
                if isinstance(v, dict):
                    found = find_candidate(v)
                    if found:
                        return found
                if isinstance(v, list):
                    for item in v:
                        if isinstance(item, dict):
                            found = find_candidate(item)
                            if found:
                                return found
                        elif isinstance(item, str) and (item.strip().startswith("files/") or item.strip().endswith(".jsonl")):
                            return item.strip()
        return None

    # Prefer explicit result.file in dump
    if isinstance(dump, dict) and "result" in dump and isinstance(dump["result"], dict):
        rf = dump["result"].get("file") or dump["result"].get("uri") or dump["result"].get("name")
        if rf and not rf.startswith("batches/"):
            return rf

    return find_candidate(dump)

# ==============================
# Robust Download with Retry
# ==============================
def download_with_retry(file_name, max_retries=5):
    for attempt in range(1, max_retries + 1):
        try:
            return client.files.download(file=file_name)

        except Exception as e:
            err_str = str(e)

            # Retry only for SSL / connection issues
            if "DECRYPTION_FAILED_OR_BAD_RECORD_MAC" in err_str or "SSL" in err_str or "Connection" in err_str:
                wait_time = 2 ** attempt
                print(f"\n    ⚠️ Download failed (attempt {attempt}/{max_retries})")
                print(f"    Reason: {e}")
                print(f"    Retrying in {wait_time}s...\n")
                time.sleep(wait_time)
                continue
            else:
                raise e

    raise Exception("Max retries reached for download")

# ==============================
# Download
# ==============================
def download():
    tracking = load_tracking()

    if not tracking:
        print("\n" + "=" * 60)
        print("  No batches found in tracking.")
        print("  Run batch_submit.py first to submit batches.")
        print("=" * 60 + "\n")
        return

    # Open workbook once
    while True:
        try:
            workbook = load_workbook(config.WORKBOOK_PATH)
            ws = workbook[config.SHEET_NAME]
            break
        except PermissionError:
            print("\nExcel file is open!")
            input("Close Excel and press ENTER to retry...")

    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip().lower()] = col

    if "imagename" not in headers or "status" not in headers:
        print("Required columns (imagename, status) not found in Excel.")
        return

    # Build imagename -> row lookup dict ONCE (huge speedup)
    print("  Building Excel lookup index...")
    imagename_to_row = {}
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=headers["imagename"]).value
        if cell:
            key = str(cell).strip().lower()
            imagename_to_row[key] = row
    print(f"  Indexed {len(imagename_to_row)} rows\n")

    # Counters
    batches_downloaded = 0
    batches_skipped = 0
    batches_not_ready = 0
    batches_failed = 0
    total_images_saved = 0
    total_images_error = 0
    total_images_no_data = 0

    print("\n" + "=" * 60)
    print("  Batch Download")
    print("=" * 60 + "\n")

    for entry in tracking:
        batch_num = entry.get("batch", entry.get("batch_number", "?"))
        entry.setdefault("downloaded", False)
        entry.setdefault("result_file", None)

        if entry["downloaded"]:
            print(f"  Batch {batch_num}: already downloaded (skipped)")
            batches_skipped += 1
            continue

        job_id = entry.get("job_id")
        if not job_id:
            print(f"  Batch {batch_num}: no job ID (skipped)")
            batches_not_ready += 1
            continue

        # Refresh status
        try:
            job = client.batches.get(name=job_id)
            entry["status"] = job.state.name
        except Exception as e:
            print(f"  Batch {batch_num}: failed to query -> {e}")
            batches_failed += 1
            continue

        if entry["status"] != "JOB_STATE_SUCCEEDED":
            label = entry["status"].replace("JOB_STATE_", "")
            print(f"  Batch {batch_num}: {label} (not ready)")
            batches_not_ready += 1
            continue

        entry["completed_at"] = entry.get("completed_at") or datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Get result file
        result_file = entry.get("result_file") or find_result_file(job)
        if not result_file:
            print(f"  Batch {batch_num}: no result file found!")
            batches_failed += 1
            continue

        entry["result_file"] = result_file

        # Download result JSONL (can be large, may take time)
        print(f"  Batch {batch_num}: downloading result file...")
        print(f"    File: {result_file}")
        
        # Start spinner in background thread
        stop_spinner = threading.Event()
        spinner_thread = threading.Thread(target=show_spinner, args=(stop_spinner, "    Downloading"))
        spinner_thread.start()
        
        try:
            content = download_with_retry(result_file)
            stop_spinner.set()
            spinner_thread.join()
            size_mb = len(content) / (1024 * 1024)
            print(f"    ✓ Downloaded {size_mb:.1f} MB")
        except Exception as e:
            stop_spinner.set()
            spinner_thread.join()
            print(f"\n  Batch {batch_num}: download FAILED -> {e}")
            batches_failed += 1
            continue

        print(f"  Batch {batch_num}: parsing JSONL...", flush=True)
        lines = content.decode("utf-8").splitlines()
        items = []
        for line in lines:
            if line.strip():
                items.append(json.loads(line))

        print(f"  Batch {batch_num}: processing {len(items)} images in parallel...")
        
        # Process images in parallel
        batch_saved = 0
        batch_error = 0
        batch_no_data = 0
        total_items = len(items)
        processed = 0

        try:
            with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
                futures = {
                    pool.submit(process_result_line, item, ws, headers, imagename_to_row): item.get("key", "?")
                    for item in items
                }
                for future in as_completed(futures):
                    key = futures[future]
                    processed += 1
                    try:
                        image_key, status = future.result()
                    except Exception as e:
                        print(f"    [{processed}/{total_items}] {key} -> ERROR: {e}")
                        batch_error += 1
                        continue

                    if status == "saved":
                        batch_saved += 1
                        print(f"    [{processed}/{total_items}] {image_key} -> saved ✓")
                    elif status == "api_error":
                        batch_error += 1
                        print(f"    [{processed}/{total_items}] {image_key or key} -> API error")
                    elif status in ("no_candidates", "no_image", "no_key"):
                        batch_no_data += 1
                        print(f"    [{processed}/{total_items}] {image_key or key} -> no data")
        
        except KeyboardInterrupt:
            print(f"\n  Batch {batch_num}: INTERRUPTED by user")
            print(f"  Progress: {processed}/{total_items} processed, {batch_saved} saved")
            print(f"  Saving workbook & tracking before exit...\n")
            workbook.save(config.WORKBOOK_PATH)
            save_tracking(tracking)
            print("  Saved. Re-run to resume from next batch.\n")
            return

        total_images_saved += batch_saved
        total_images_error += batch_error
        total_images_no_data += batch_no_data

        entry["downloaded"] = True
        print(f"  Batch {batch_num}: done -> {batch_saved} saved, {batch_error} errors, {batch_no_data} no data\n")
        batches_downloaded += 1

        # Save workbook after each batch
        while True:
            try:
                workbook.save(config.WORKBOOK_PATH)
                break
            except PermissionError:
                print("\n  Excel file is open!")
                input("  Close Excel and press ENTER to retry...")

        # Save tracking after each batch (resume-safe)
        save_tracking(tracking)

    save_tracking(tracking)

    # ── Cleanup if all done ──────────────────────────────────────
    all_done = tracking and all(e.get("downloaded") for e in tracking)
    if all_done:
        save_tracking([])
        if os.path.exists("batches"):
            shutil.rmtree("batches")
        
        if os.path.exists("uploaded_files.json"):
            os.remove("uploaded_files.json")

    # ── Summary ──────────────────────────────────────────────────
    total_batches = len(tracking)
    print(f"\n{'=' * 60}")
    print(f"  Download Summary")
    print(f"{'=' * 60}")
    print(f"  Total batches          : {total_batches}")
    print(f"  Downloaded this run    : {batches_downloaded}")
    print(f"  Already downloaded     : {batches_skipped}")
    print(f"  Not ready (running)    : {batches_not_ready}")
    print(f"  Failed                 : {batches_failed}")
    print(f"  ---")
    print(f"  Images saved           : {total_images_saved}")
    print(f"  Images with errors     : {total_images_error}")
    print(f"  Images with no data    : {total_images_no_data}")

    if all_done:
        print(f"\n  All batches downloaded! Tracking & batch files cleaned up.")
    elif batches_not_ready > 0:
        print(f"\n  {batches_not_ready} batch(es) still running.")
        print(f"  Run batch_monitor.py to check status, then re-run this.")

    print(f"{'=' * 60}\n")


if __name__ == "__main__":
    download()
