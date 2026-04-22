import os
import json
import requests
import time
import random
from datetime import datetime
from openpyxl import load_workbook
from google import genai
import config
from concurrent.futures import ThreadPoolExecutor, as_completed
from io import BytesIO
import hashlib
import threading

# ==============================
# SETTINGS
# ==============================
BATCH_FOLDER = "batches"
TRACK_FILE = "batch_tracking.json"
UPLOAD_CACHE = "uploaded_files.json"
MAX_WORKERS = 50
MAX_RETRIES = 3

os.makedirs(BATCH_FOLDER, exist_ok=True)
client = genai.Client(api_key=config.GEMINI_API_KEY)

# Thread-safe locks
cache_lock = threading.Lock()
save_lock = threading.Lock()
wb_lock = threading.Lock()

def check_batch_count():
    if config.TOTAL_BATCHES == 0:
        print("\n❌ TOTAL_BATCHES = 0")
        print("👉 Please set batch count in config.py\n")
        return False
    return True

# ==============================
# Helpers
# ==============================
def load_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default


def save_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4)


def update_config_total_batches(value):
    """Overwrite TOTAL_BATCHES in config.py."""
    path = "config.py"
    if not os.path.exists(path):
        return
    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    with open(path, "w", encoding="utf-8") as f:
        for line in lines:
            if line.strip().startswith("TOTAL_BATCHES"):
                f.write(f"TOTAL_BATCHES = {value}   # Number of batches to submit at once\n")
            else:
                f.write(line)


def is_retryable_quota_error(exc):
    """Return True for transient rate-limit/quota style failures."""
    msg = str(exc).lower()
    retry_markers = (
        "429",
        "resource_exhausted",
        "rate limit",
        "too many requests",
        "quota",
        "try again later",
    )
    return any(marker in msg for marker in retry_markers)


def call_with_backoff(action, fn, max_attempts=6, base_delay=8, max_delay=180):
    """Call fn() with exponential backoff for retryable quota/rate errors."""
    for attempt in range(1, max_attempts + 1):
        try:
            return fn()
        except Exception as e:
            retryable = is_retryable_quota_error(e)
            if (not retryable) or attempt == max_attempts:
                raise

            delay = min(max_delay, base_delay * (2 ** (attempt - 1)))
            # Small jitter avoids synchronized retries when many jobs run at once.
            wait = delay + random.uniform(0.0, 1.5)
            print(
                f"  {action}: retryable error on attempt {attempt}/{max_attempts} -> {e}"
            )
            print(f"  {action}: waiting {wait:.1f}s before retry...")
            time.sleep(wait)


# ==============================
# Worker: Download + Upload one image
# ==============================
def download_and_upload(row, image_url, prompt, image_name, upload_cache,
                        workbook, ws, status_col):
    """
    Downloads one image, uploads to Gemini, saves cache & Excel per-image.
    Returns (image_name, entry_or_None, row, status_str).
    """
    # Already cached? Fix Excel if needed and return
    with cache_lock:
        if image_name in upload_cache:
            with wb_lock:
                ws.cell(row=row, column=status_col).value = "uploaded"
            return image_name, upload_cache[image_name], row, "cached"

    # --- Download ---
    image_bytes = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.get(image_url, timeout=30)
            resp.raise_for_status()
            image_bytes = resp.content
            break
        except Exception:
            if attempt < MAX_RETRIES:
                time.sleep(2 * attempt)

    if image_bytes is None:
        return image_name, None, row, "download_failed"

    # --- Upload to Gemini ---
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            uploaded = client.files.upload(
                file=BytesIO(image_bytes),
                config={"mime_type": "image/jpeg"},
            )
            file_id = uploaded.name.split("/")[-1]
            file_uri = (
                f"https://generativelanguage.googleapis.com/v1beta/files/{file_id}"
            )

            entry = {
                "key": image_name,
                "request": {
                    "contents": [
                        {
                            "parts": [
                                {"text": prompt},
                                {"file_data": {"file_uri": file_uri}},
                            ]
                        }
                    ],
                    "generation_config": {"response_modalities": ["IMAGE"]},
                },
            }

            # --- Save immediately (zero data loss) ---
            with cache_lock:
                upload_cache[image_name] = entry
            with save_lock:
                save_json(UPLOAD_CACHE, upload_cache)

            # Mark row as "uploaded" right away
            with wb_lock:
                ws.cell(row=row, column=status_col).value = "uploaded"

            return image_name, entry, row, "uploaded"

        except Exception:
            if attempt < MAX_RETRIES:
                time.sleep(2 * attempt)

    return image_name, None, row, "upload_failed"


# ==============================
# Main
# ==============================
def main():
    if not check_batch_count():
        return

    # ── Step 0: Load everything ──────────────────────────────────────
    workbook = load_workbook(config.WORKBOOK_PATH)
    ws = workbook[config.SHEET_NAME]

    headers = {}
    for col in range(1, ws.max_column + 1):
        h = ws.cell(row=1, column=col).value
        if h:
            headers[str(h).strip().lower()] = col

    for r in ("imageurl", "prompt", "imagename", "status"):
        if r not in headers:
            raise RuntimeError(f"Missing column: {r}")

    status_col = headers["status"]
    upload_cache = load_json(UPLOAD_CACHE, {})
    tracking = load_json(TRACK_FILE, [])

    # ── Step 1: Classify rows ────────────────────────────────────────
    # pending  → needs download+upload
    # uploaded → already uploaded, ready for batching
    # batched  → done, skip entirely
    rows_pending = []   # need download + upload
    rows_uploaded = []  # already uploaded (in cache), ready for batch

    # First pass: fix mismatches (cached but Excel still says pending)
    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row=row, column=status_col).value
        status = str(raw).strip().lower() if raw else ""
        if status == "batched":
            continue

        image_url = ws.cell(row=row, column=headers["imageurl"]).value
        prompt = ws.cell(row=row, column=headers["prompt"]).value
        image_name = ws.cell(row=row, column=headers["imagename"]).value
        if not (image_url and prompt and image_name):
            continue

        image_name = str(image_name).strip()

        if image_name in upload_cache:
            # Cached → ensure Excel says "uploaded"
            if status != "uploaded":
                ws.cell(row=row, column=status_col).value = "uploaded"
            rows_uploaded.append((row, image_url, prompt, image_name))
        else:
            # Not cached → needs processing
            if status not in ("pending", "uploaded", "staged"):
                continue
            rows_pending.append((row, image_url, prompt, image_name))

    # save any mismatch fixes
    workbook.save(config.WORKBOOK_PATH)

    total_uploaded = len(rows_uploaded)
    total_pending = len(rows_pending)
    total = total_uploaded + total_pending

    print(f"\n{'='*60}")
    print(f"  Image Status Summary")
    print(f"{'='*60}")
    print(f"  Already uploaded (cached) : {total_uploaded}")
    print(f"  Need download + upload    : {total_pending}")
    print(f"  Total to batch            : {total}")
    print(f"{'='*60}\n")

    if total == 0:
        print("Nothing to process.")
        return

    # ── Step 2: Download + Upload pipeline ───────────────────────────
    if total_pending > 0:
        print(f"Processing {total_pending} images (download + upload)...\n")

        done = 0
        ok = 0
        fail = 0
        wb_save_counter = 0

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            futures = {
                pool.submit(
                    download_and_upload,
                    row, img, pr, name, upload_cache,
                    workbook, ws, status_col,
                ): name
                for row, img, pr, name in rows_pending
            }

            for future in as_completed(futures):
                name = futures[future]
                try:
                    img_name, result, row_num, status = future.result()
                except Exception:
                    img_name, result, row_num, status = name, None, 0, "error"

                done += 1

                if status in ("uploaded", "cached"):
                    ok += 1
                    if result:
                        rows_uploaded.append((row_num, "", "", img_name))
                    print(f"  [{done}/{total_pending}] {img_name} -> uploaded")
                elif status == "download_failed":
                    fail += 1
                    print(f"  [{done}/{total_pending}] {img_name} -> DOWNLOAD FAILED")
                elif status == "upload_failed":
                    fail += 1
                    print(f"  [{done}/{total_pending}] {img_name} -> UPLOAD FAILED")
                else:
                    fail += 1
                    print(f"  [{done}/{total_pending}] {img_name} -> ERROR")

                # Save workbook every 50 images
                wb_save_counter += 1
                if wb_save_counter >= 50:
                    workbook.save(config.WORKBOOK_PATH)
                    wb_save_counter = 0

        # Final save after pipeline
        workbook.save(config.WORKBOOK_PATH)
        save_json(UPLOAD_CACHE, upload_cache)

        print(f"\n  Pipeline done: {ok} uploaded, {fail} failed\n")
    else:
        print("All images already uploaded (from cache).\n")

    # ── Step 3: Batch creation ───────────────────────────────────────
    # Collect all "uploaded" rows for batching
    all_requests = []
    for row in range(2, ws.max_row + 1):
        raw = ws.cell(row=row, column=status_col).value
        status = str(raw).strip().lower() if raw else ""
        if status != "uploaded":
            continue
        image_name = str(ws.cell(row=row, column=headers["imagename"]).value).strip()
        if image_name in upload_cache:
            all_requests.append((upload_cache[image_name], row))

    if not all_requests:
        print("No uploaded rows to batch.")
        return

    # Sort for consistent batch composition
    all_requests.sort(key=lambda x: x[0].get("key", ""))

    total_rows = len(all_requests)
    actual_batches = min(config.TOTAL_BATCHES, total_rows)
    if actual_batches <= 0:
        print("TOTAL_BATCHES is 0 — nothing to submit.")
        return

    base_size = total_rows // actual_batches
    remainder = total_rows % actual_batches

    # Sequential numbering from highest existing batch
    highest = max((t.get("batch", 0) for t in tracking), default=0) if tracking else 0

    print(f"{'='*60}")
    print(f"  Batch Creation")
    print(f"{'='*60}")
    print(f"  Rows to batch   : {total_rows}")
    print(f"  Batches to make : {actual_batches}")
    print(f"  Next batch #    : {highest + 1}")
    print(f"{'='*60}\n")

    idx = 0
    created = 0
    skipped = 0
    failed_batches = 0

    for i in range(actual_batches):
        batch_num = highest + i + 1
        # distribute remainder evenly: first `remainder` batches get +1 row
        size = base_size + (1 if i < remainder else 0)
        batch_items = all_requests[idx:idx + size]
        idx += size

        batch_reqs = [item[0] for item in batch_items]
        batch_rows = [item[1] for item in batch_items]

        # Stable signature for dedup
        batch_keys = sorted([r.get("key", "") for r in batch_reqs])
        sig = hashlib.sha256("|".join(batch_keys).encode()).hexdigest()

        # Check if already submitted
        existing = next((t for t in tracking if t.get("batch_sig") == sig), None)
        if existing and existing.get("job_id"):
            print(f"  Batch {batch_num}: already submitted (skipped)")
            for rn in batch_rows:
                ws.cell(row=rn, column=status_col).value = "batched"
            workbook.save(config.WORKBOOK_PATH)
            skipped += 1
            continue

        # Write JSONL
        batch_file = os.path.join(BATCH_FOLDER, f"batch_{batch_num}.jsonl")
        with open(batch_file, "w", encoding="utf-8") as f:
            for r in batch_reqs:
                f.write(json.dumps(r) + "\n")

        # Upload batch file
        try:
            print(f"  Batch {batch_num}: uploading JSONL ({len(batch_reqs)} rows)...")
            uploaded_batch = client.files.upload(
                file=batch_file,
                config={"mime_type": "text/plain"},
            )
            file_src = uploaded_batch.name
        except Exception as e:
            print(f"  Batch {batch_num}: JSONL upload FAILED — {e}")
            print(f"  Stopping. Re-run to retry from this batch.\n")
            failed_batches += 1
            break

        # Save tracking immediately after file upload (crash-safe)
        if existing:
            existing.update({"file_src": file_src, "batch": batch_num})
        else:
            tracking.append({
                "batch_sig": sig,
                "batch_keys": batch_keys,
                "file_src": file_src,
                "batch": batch_num,
                "rows": len(batch_reqs),
                "status": "file_uploaded",
                "job_id": None,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "completed_at": None,
            })
        save_json(TRACK_FILE, tracking)

        # Create batch job
        try:
            print(f"  Batch {batch_num}: creating job...")
            job = call_with_backoff(
                action=f"Batch {batch_num}: creating job",
                fn=lambda: client.batches.create(
                    model=config.MODEL_ID,
                    src=file_src,
                    config={"display_name": f"Batch_{batch_num}"},
                ),
            )
        except Exception as e:
            print(f"  Batch {batch_num}: job creation FAILED — {e}")
            print(f"  The JSONL file is saved. Re-run to retry.\n")
            failed_batches += 1
            break

        # Update tracking with job info
        entry = next((t for t in tracking if t.get("batch_sig") == sig), None)
        if entry:
            entry.update({
                "job_id": job.name,
                "status": job.state.name,
                "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            })
        save_json(TRACK_FILE, tracking)
        
        time.sleep(5)
        
        # Mark rows as "batched"
        for rn in batch_rows:
            ws.cell(row=rn, column=status_col).value = "batched"
        workbook.save(config.WORKBOOK_PATH)

        # Decrement TOTAL_BATCHES in config
        new_total = max(0, config.TOTAL_BATCHES - 1)
        update_config_total_batches(new_total)
        config.TOTAL_BATCHES = new_total

        created += 1
        print(f"  Batch {batch_num}: SUBMITTED -> {job.name}")
        print(f"  (config TOTAL_BATCHES now = {new_total})\n")

    # ── Summary ──────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"  Summary")
    print(f"{'='*60}")
    print(f"  Batches created  : {created}")
    print(f"  Batches skipped  : {skipped}")
    print(f"  Batches failed   : {failed_batches}")
    print(f"  TOTAL_BATCHES    : {config.TOTAL_BATCHES}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()

